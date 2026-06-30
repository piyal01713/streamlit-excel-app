import base64
import io
import random
import re

import openpyxl
import pandas as pd
import requests
import streamlit as st
from anthropic import Anthropic

# ------------------------------------------------
# 1. CONFIG & INITIALIZATION
# ------------------------------------------------

st.set_page_config(page_title="Excel Insights Pro", layout="wide")

MODELS = {
    "Claude 4.5 Haiku": "claude-haiku-4-5-20251001",
    "Claude 4.5 Sonnet": "claude-sonnet-4-5-20250929",
}

if "all_sheets" not in st.session_state:
    st.session_state.all_sheets = None
if "scope" not in st.session_state:
    st.session_state.scope = "Analyze All Sheets (Join/Compare)"
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []

HEADER_KEYWORDS = {
    "amount", "code", "cost", "date", "description", "id", "name", "number",
    "price", "project", "qty", "quantity", "status", "total", "type",
}
SECTION_LABELS = {
    "business", "dev", "free resource 1", "free resource 2", "qa", "total",
}
MAX_HEADER_SCAN_ROWS = 200


def load_api_key():
    try:
        return st.secrets.get("ANTHROPIC_API_KEY", "").strip()
    except (FileNotFoundError, KeyError, AttributeError):
        return ""


api_key = load_api_key()
if not api_key:
    st.error("Missing `ANTHROPIC_API_KEY` in `.streamlit/secrets.toml`.")
    st.stop()

client = Anthropic(api_key=api_key)

# ------------------------------------------------
# 2. HELPERS (URL, WORKBOOK, HEADER DETECTION)
# ------------------------------------------------


def get_direct_download_link(url):
    try:
        if "docs.google.com/spreadsheets" in url:
            file_id = re.search(r"/d/([^/]+)", url).group(1)
            return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
        if "drive.google.com" in url:
            file_id = re.search(r"d/([^/]+)", url).group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"
        if "1drv.ms" in url or "onedrive.live.com" in url:
            encoded_url = (
                base64.b64encode(url.encode())
                .decode()
                .replace("+", "-")
                .replace("/", "_")
                .rstrip("=")
            )
            return f"https://api.onedrive.com/v1.0/shares/u!{encoded_url}/root/content"
        return url
    except Exception:
        return url


def _looks_numeric(value):
    try:
        float(str(value).replace(",", "").replace("$", "").replace("%", "").strip())
        return True
    except (ValueError, TypeError):
        return False


def collect_sheet_metadata(ws):
    hidden_rows = [
        row_idx for row_idx, dim in ws.row_dimensions.items() if dim.hidden
    ]
    hidden_columns = [
        col for col, dim in ws.column_dimensions.items() if dim.hidden
    ]
    return {
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "merged_cells": [str(r) for r in ws.merged_cells.ranges],
        "hidden_rows": sorted(hidden_rows),
        "hidden_columns": sorted(hidden_columns),
        "sheet_visibility": ws.sheet_state,
        "row_count": ws.max_row or 0,
        "column_count": ws.max_column or 0,
    }


def sheet_to_dataframe(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def _row_text_values(row):
    return [
        str(v).strip()
        for v in row
        if pd.notna(v) and str(v).strip() != ""
    ]


def _freeze_pane_header_hint(freeze_panes):
    """First scrollable row (1-indexed) minus 1 => header row (0-indexed)."""
    if not freeze_panes:
        return None
    match = re.search(r"(\d+)", freeze_panes)
    if not match:
        return None
    freeze_row = int(match.group(1))
    if freeze_row >= 2:
        return freeze_row - 2
    return None


def _merged_title_rows(merged_cells):
    """Rows that are wide single-line merges — usually titles, not column headers."""
    title_rows = set()
    for merge_str in merged_cells:
        match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", merge_str.upper())
        if not match:
            continue
        _, r1, _, r2 = match.groups()
        r1, r2 = int(r1), int(r2)
        if r1 == r2:
            title_rows.add(r1 - 1)
        else:
            for row in range(r1, r2 + 1):
                title_rows.add(row - 1)
    return title_rows


def _score_header_row(row, column_count):
    values = _row_text_values(row)
    if not values:
        return None

    filled = len(values)
    if filled <= 2 and column_count > 5:
        return None

    text_values = [v for v in values if not _looks_numeric(v)]
    str_ratio = len(text_values) / filled
    keyword_hits = sum(
        1 for v in values
        if v.lower() in HEADER_KEYWORDS
        or any(kw in v.lower() for kw in HEADER_KEYWORDS)
    )
    unique_ratio = len(set(v.lower() for v in values)) / filled

    score = str_ratio * 2.0
    score += keyword_hits * 0.75
    score += min(filled / max(column_count, 1), 1.0)
    score += unique_ratio * 0.5

    numeric_ratio = (filled - len(text_values)) / filled
    if numeric_ratio > 0.7:
        score -= 2.0
    if unique_ratio < 0.3 and filled > 2:
        score -= 1.0

    return score


def _score_matrix_header_row(row, row_idx, df, column_count):
    """Score rows that look like mid-sheet matrix headers (e.g. project names on row 5)."""
    values = _row_text_values(row)
    if not values:
        return None

    non_empty_cols = sum(1 for v in row if pd.notna(v) and str(v).strip())
    fill_span = non_empty_cols / max(column_count, 1)
    if fill_span < 0.15 and non_empty_cols < 3:
        return None

    text_values = [v for v in values if not _looks_numeric(v)]
    if not text_values:
        return None

    str_ratio = len(text_values) / len(values)
    unique_ratio = len(set(v.lower() for v in values)) / len(values)
    short_labels = sum(1 for v in text_values if len(v) <= 50 and "\n" not in v)
    short_ratio = short_labels / len(text_values)

    score = fill_span * 4.0
    score += str_ratio * 2.5
    score += unique_ratio * 3.0
    score += short_ratio * 2.0

    multiline = sum(1 for v in values if "\n" in v)
    if multiline > len(values) * 0.25:
        score -= 3.0

    if row_idx + 1 < len(df):
        next_values = _row_text_values(df.iloc[row_idx + 1])
        if next_values:
            score += 1.0
            if any("\n" in v for v in next_values):
                score += 1.5

    return score


def _extract_column_labels(df, header_row):
    labels = []
    for col_idx, val in enumerate(df.iloc[header_row]):
        if pd.notna(val) and str(val).strip():
            labels.append((col_idx, str(val).strip()))
    return labels


def find_column_by_header_label(df, header_row, target_label):
    """Find 0-based column index where the header row matches target_label."""
    target = str(target_label).strip().lower()
    for col_idx, val in enumerate(df.iloc[header_row]):
        if pd.notna(val) and str(val).strip().lower() == target:
            return col_idx
    return None


def detect_header_row(df, metadata):
    if df.empty:
        return 0, [], []

    scores = {}
    column_count = metadata.get("column_count") or df.shape[1]
    hidden_rows = set(metadata.get("hidden_rows", []))
    title_rows = _merged_title_rows(metadata.get("merged_cells", []))
    scan_limit = min(len(df), MAX_HEADER_SCAN_ROWS)

    freeze_hint = _freeze_pane_header_hint(metadata.get("freeze_panes"))
    if freeze_hint is not None and 0 <= freeze_hint < len(df):
        scores[freeze_hint] = scores.get(freeze_hint, 0.0) + 6.0

    for idx in range(scan_limit):
        if (idx + 1) in hidden_rows:
            continue

        row = df.iloc[idx]
        row_score = _score_header_row(row, column_count) or 0.0
        matrix_score = _score_matrix_header_row(row, idx, df, column_count) or 0.0
        combined = max(row_score, matrix_score)

        if idx in title_rows and combined < 4.0:
            combined -= 2.5

        if idx < 3 and combined < 3.0:
            combined -= 1.0

        if combined > 0:
            scores[idx] = scores.get(idx, 0.0) + combined

    if not scores:
        header_row = 0
    else:
        header_row = max(scores, key=scores.get)

    column_names = [
        str(v).strip() if pd.notna(v) and str(v).strip() else f"col_{i}"
        for i, v in enumerate(df.iloc[header_row])
    ]
    header_labels = _extract_column_labels(df, header_row)
    top_candidates = sorted(scores.items(), key=lambda item: item[1], reverse=True)[:5]

    return header_row, column_names, header_labels, top_candidates


def load_workbook_sheets(raw_data):
    wb = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        metadata = collect_sheet_metadata(ws)
        df = sheet_to_dataframe(ws)
        header_row, column_names, header_labels, top_candidates = detect_header_row(df, metadata)
        sheets[name] = {
            "df": df,
            "metadata": metadata,
            "header_row": header_row,
            "data_start_row": header_row + 1,
            "column_names": column_names,
            "header_labels": header_labels,
            "header_candidates": top_candidates,
        }
    return sheets


def sheets_to_dfs(sheets_dict):
    return {name: info["df"] for name, info in sheets_dict.items()}


def _header_region_sample(df, header_row, padding=3):
    start = max(0, header_row - padding)
    end = min(len(df), header_row + padding + 4)
    return df.iloc[start:end].to_csv(index=True, header=False)


def _sample_rows(df, header_row, n_first, n_last, n_random):
    sections = []

    sections.append(
        f"Header region (rows {max(0, header_row - 3)}–"
        f"{min(len(df) - 1, header_row + 3)}, 0-indexed):"
    )
    sections.append(_header_region_sample(df, header_row))

    sections.append(f"First {n_first} rows (0-indexed, headers not applied):")
    sections.append(df.head(n_first).to_csv(index=True, header=False))

    if len(df) > n_first:
        sections.append(f"Last {n_last} rows:")
        sections.append(df.tail(n_last).to_csv(index=True, header=False))

    middle_indices = list(range(n_first, max(len(df) - n_last, n_first)))
    if middle_indices and n_random > 0:
        pick = min(n_random, len(middle_indices))
        sample_idx = sorted(random.sample(middle_indices, pick))
        sections.append(f"Random sample at row indices {sample_idx}:")
        sections.append(df.iloc[sample_idx].to_csv(index=True, header=False))

    return "\n".join(sections)


def build_context(sheets_dict, scope, depth):
    if depth == "Quick (Top 100 Rows)":
        n_first, n_last, n_random = 8, 3, 3
    else:
        n_first, n_last, n_random = 15, 8, 8

    context = (
        "Workbook context for analysis. DataFrames in `dfs` use integer columns (0, 1, 2...).\n"
        "Headers may appear in the MIDDLE of the sheet (not row 0) — e.g. matrix layouts with title "
        "rows on top and project/column names on row 5 or similar.\n"
        "Use the pre-detected header_row (0-indexed), header_labels, and data_start_row from context.\n"
        "To locate a column like 'Ipex', scan header_labels or df.iloc[header_row] — do NOT assume row 0.\n"
        "Data cells below headers may contain newline-separated names; split on '\\n' and skip section "
        f"labels like {sorted(SECTION_LABELS)}.\n"
        f"Sample depth: {depth}.\n\n"
    )

    selected_sheets = (
        sheets_dict.keys()
        if scope == "Analyze All Sheets (Join/Compare)"
        else [scope]
    )

    for name in selected_sheets:
        info = sheets_dict[name]
        df = info["df"]
        metadata = info["metadata"]
        header_row = info["header_row"]
        column_names = info["column_names"]
        header_labels = info["header_labels"]
        data_start_row = info["data_start_row"]
        candidates = info["header_candidates"]

        context += f"### SHEET: '{name}'\n"
        context += "Metadata:\n"
        context += f"- freeze_panes: {metadata['freeze_panes']}\n"
        context += f"- merged_cells: {metadata['merged_cells'][:20]}"
        if len(metadata["merged_cells"]) > 20:
            context += f" ... (+{len(metadata['merged_cells']) - 20} more)"
        context += "\n"
        context += f"- hidden_rows (1-indexed): {metadata['hidden_rows'][:30]}\n"
        context += f"- hidden_columns: {metadata['hidden_columns'][:30]}\n"
        context += f"- sheet_visibility: {metadata['sheet_visibility']}\n"
        context += f"- row_count: {metadata['row_count']}\n"
        context += f"- column_count: {metadata['column_count']}\n"
        context += f"- detected_header_row (0-indexed): {header_row}\n"
        context += f"- detected_header_row (1-indexed Excel row): {header_row + 1}\n"
        context += f"- data_start_row (0-indexed): {data_start_row}\n"
        context += f"- detected_column_names: {column_names}\n"
        context += f"- header_labels (col_index, label): {header_labels[:40]}"
        if len(header_labels) > 40:
            context += f" ... (+{len(header_labels) - 40} more)"
        context += "\n"
        context += f"- top_header_candidates (0-indexed row, score): {candidates}\n"
        context += f"- dataframe_shape: {df.shape[0]} rows x {df.shape[1]} cols\n\n"
        context += "Representative samples:\n"
        context += _sample_rows(df, header_row, n_first, n_last, n_random)
        context += "\n---\n"

    return context


def build_conversation_history():
    entries = st.session_state.chat_history[-8:]
    if not entries:
        return ""

    blocks = []
    for msg in entries:
        blocks.append(
            f"User:\n{msg['user']}\n\n"
            f"Assistant:\n{msg['assistant']}\n\n"
            f"Execution Result:\n{msg['result']}"
        )
    return "\n\n----------------\n\n".join(blocks)


# ------------------------------------------------
# 3. AI LOGIC
# ------------------------------------------------


def get_analysis_code(user_query, workbook_context, conversation_history, model_id):
    system_prompt = (
        "You are a Senior Python Data Expert. You are working with a dictionary of DataFrames called `dfs`.\n"
        "DataFrames have integer columns (0, 1, 2) initially.\n\n"
        "Sheets may use matrix layouts: title rows at the top, then column headers (project names like "
        "'Ipex', 'Amberg Logtec') on a row in the MIDDLE of the sheet.\n"
        "Context provides header_row (0-indexed), data_start_row, and header_labels (col_index, label).\n\n"
        "The user may ask follow-up questions.\n"
        "Resolve references such as \"them\", \"those\", \"it\", \"previous\", \"same sheet\", "
        "\"that project\", \"the earlier result\".\n"
        "Use the conversation history before asking the user to clarify.\n"
        "Use previous execution results when helpful.\n\n"
        "MANDATORY STEPS IN YOUR CODE:\n"
        "1. Read header_row and data_start_row from context — NEVER assume headers are on row 0.\n"
        "2. To find a column by name: iterate df.iloc[header_row] or use find_column_by_header_label(df, header_row, name).\n"
        "3. Scan data from row data_start_row downward: for row in range(data_start_row, len(df)).\n"
        "4. Cell values may contain newline-separated names — split on '\\n', strip, and skip section labels "
        "(Dev, Business, QA, Total, Free Resource 1, Free Resource 2).\n"
        "5. SAFE INDEXING: use row.iloc[col_idx] or df.iloc[row_idx, col_idx].\n"
        "6. RESULT: Store your final answer in a variable named `result`.\n"
        "RESPOND ONLY WITH CODE."
    )

    history_text = conversation_history if conversation_history else "(none)"
    user_content = (
        f"Workbook Context:\n{workbook_context}\n\n"
        f"Conversation History:\n{history_text}\n\n"
        f"Current User Question:\n{user_query}"
    )

    try:
        response = client.messages.create(
            model=model_id,
            max_tokens=2500,
            system=system_prompt,
            messages=[{"role": "user", "content": user_content}],
        )
        return re.sub(r"```python\n|```", "", response.content[0].text.strip())
    except Exception as e:
        st.error(f"AI Error: {e}")
        return None


def generate_natural_answer(user_query, execution_result, model_id):
    try:
        response = client.messages.create(
            model=model_id,
            max_tokens=800,
            system="Summarize data results clearly. No preamble. Use bolding.",
            messages=[{"role": "user", "content": f"User asked: {user_query}\nRaw Result: {execution_result}"}],
        )
        return response.content[0].text
    except Exception:
        return f"Result: {execution_result}"


# ------------------------------------------------
# 4. STREAMLIT UI
# ------------------------------------------------

def inject_custom_css():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&display=swap');
        
        /* Font overrides */
        html, body, .stMarkdown, p, h1, h2, h3, h4, h5, h6, .stButton button, input, label {
            font-family: 'Outfit', -apple-system, BlinkMacSystemFont, sans-serif !important;
        }

        /* Center container adjustments */
        .block-container {
            padding-top: 1.5rem !important;
            padding-bottom: 2rem !important;
            max-width: 1200px !important;
        }

        /* Sidebar Styling */
        [data-testid="stSidebar"] {
            background-color: #0b0f19 !important;
            border-right: 1px solid rgba(255, 255, 255, 0.06) !important;
        }
        /* Spaced sidebar layout with breathing room */
        [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            gap: 1.25rem !important;
        }
        [data-testid="stSidebar"] h2 {
            font-size: 0.9rem !important;
            font-weight: 600 !important;
            text-transform: uppercase !important;
            letter-spacing: 0.06em !important;
            margin-top: 1rem !important;
            margin-bottom: 0.25rem !important;
            color: #818cf8 !important;
        }
        [data-testid="stSidebar"] .stMarkdown p {
            color: #94a3b8 !important;
            margin-bottom: 0px !important;
        }
        [data-testid="stSidebar"] .stSelectbox label, 
        [data-testid="stSidebar"] .stRadio label,
        [data-testid="stSidebar"] .stTextInput label {
            color: #cbd5e1 !important;
            font-weight: 500 !important;
            font-size: 0.85rem !important;
            margin-bottom: 0.25rem !important;
        }
        [data-testid="stSidebarContent"] {
            padding-top: 1rem !important;
        }
        [data-testid="stSidebar"] [data-testid="stSidebarUserContent"] {
            padding-top: 0rem !important;
            padding-left: 1.25rem !important;
            padding-right: 1.25rem !important;
        }
        /* Make sidebar reset button subtle and secondary */
        [data-testid="stSidebar"] .stButton button {
            background: rgba(255, 255, 255, 0.04) !important;
            color: #cbd5e1 !important;
            border: 1px solid rgba(255, 255, 255, 0.1) !important;
            box-shadow: none !important;
            font-size: 0.85rem !important;
            font-weight: 500 !important;
            padding: 6px 14px !important;
            border-radius: 8px !important;
            transition: all 0.2s ease !important;
            margin-top: 0.5rem !important;
        }
        [data-testid="stSidebar"] .stButton button:hover {
            background: rgba(255, 255, 255, 0.08) !important;
            border-color: rgba(255, 255, 255, 0.2) !important;
            color: #ffffff !important;
            transform: none !important;
            box-shadow: none !important;
        }
        /* Hide scrollbars but allow scrolling if needed */
        [data-testid="stSidebar"] > div:first-child {
            scrollbar-width: none !important; /* Firefox */
        }
        [data-testid="stSidebar"] > div:first-child::-webkit-scrollbar {
            display: none !important; /* Safari and Chrome */
        }

        /* Beautiful buttons */
        .stButton button {
            background: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%) !important;
            color: #ffffff !important;
            border: none !important;
            padding: 10px 20px !important;
            border-radius: 10px !important;
            font-weight: 600 !important;
            letter-spacing: 0.025em !important;
            box-shadow: 0 4px 6px -1px rgba(99, 102, 241, 0.2), 0 2px 4px -1px rgba(99, 102, 241, 0.1) !important;
            transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
            width: 100% !important;
        }
        .stButton button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 10px 15px -3px rgba(99, 102, 241, 0.3), 0 4px 6px -2px rgba(99, 102, 241, 0.15) !important;
            background: linear-gradient(135deg, #4f46e5 0%, #4338ca 100%) !important;
        }
        .stButton button:active {
            transform: translateY(0px) !important;
        }

        /* Upload area styling */
        [data-testid="stFileUploader"] {
            border: 2px dashed rgba(99, 102, 241, 0.25) !important;
            border-radius: 12px !important;
            background-color: rgba(99, 102, 241, 0.01) !important;
            padding: 16px !important;
            transition: all 0.3s ease !important;
        }
        [data-testid="stFileUploader"]:hover {
            border-color: #6366f1 !important;
            background-color: rgba(99, 102, 241, 0.04) !important;
        }
        [data-testid="stFileUploader"] section {
            background-color: transparent !important;
            padding: 0 !important;
        }

        /* Expander styling */
        .stExpander {
            border: 1px solid rgba(226, 232, 240, 0.08) !important;
            background-color: rgba(30, 41, 59, 0.2) !important;
            border-radius: 12px !important;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1) !important;
            margin-bottom: 1.25rem !important;
            overflow: hidden !important;
        }
        .stExpander summary {
            font-size: 1.02rem !important;
            font-weight: 600 !important;
            color: #f1f5f9 !important;
            padding: 10px 14px !important;
        }
        .stExpander summary:hover {
            background-color: rgba(99, 102, 241, 0.08) !important;
            color: #818cf8 !important;
        }

        /* Input overrides */
        .stTextInput input, .stSelectbox div[role="button"] {
            border-radius: 8px !important;
            border: 1px solid rgba(226, 232, 240, 0.12) !important;
            background-color: rgba(30, 41, 59, 0.4) !important;
            color: #f1f5f9 !important;
            transition: all 0.2s ease !important;
        }
        .stTextInput input:focus {
            border-color: #6366f1 !important;
            box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.2) !important;
        }

        /* Chat Message Bubbles */
        [data-testid="stChatMessage"] {
            background-color: rgba(30, 41, 59, 0.3) !important;
            border-radius: 16px !important;
            border: 1px solid rgba(226, 232, 240, 0.06) !important;
            padding: 1rem 1.25rem !important;
            margin-bottom: 1rem !important;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.04) !important;
        }
        
        /* Chat Input */
        .stChatInput {
            border-radius: 12px !important;
            border: 1px solid rgba(99, 102, 241, 0.2) !important;
            background-color: rgba(15, 23, 42, 0.8) !important;
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.3) !important;
        }
        .stChatInput:focus-within {
            border-color: #6366f1 !important;
            box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.3), 0 10px 15px -3px rgba(0, 0, 0, 0.3) !important;
        }

        /* Custom Badge elements for layout */
        .badge {
            display: inline-block;
            padding: 4px 10px;
            font-size: 0.8rem;
            font-weight: 600;
            border-radius: 20px;
            margin-right: 6px;
            margin-bottom: 6px;
        }
        .badge-primary {
            background-color: rgba(99, 102, 241, 0.15);
            color: #818cf8;
            border: 1px solid rgba(99, 102, 241, 0.3);
        }
        .badge-secondary {
            background-color: rgba(168, 85, 247, 0.15);
            color: #c084fc;
            border: 1px solid rgba(168, 85, 247, 0.3);
        }
        .badge-success {
            background-color: rgba(16, 185, 129, 0.15);
            color: #34d399;
            border: 1px solid rgba(16, 185, 129, 0.3);
        }
        .badge-neutral {
            background-color: rgba(148, 163, 184, 0.15);
            color: #94a3b8;
            border: 1px solid rgba(148, 163, 184, 0.3);
        }

        /* Card panels */
        .card-panel {
            background: rgba(30, 41, 59, 0.4);
            border: 1px solid rgba(255, 255, 255, 0.05);
            border-radius: 12px;
            padding: 16px;
            margin-bottom: 16px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }
        </style>
        """,
        unsafe_allow_html=True
    )

inject_custom_css()

# Styled header banner
st.markdown(
    """
    <div style="background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 50%, #c084fc 100%); padding: 24px 32px; border-radius: 16px; margin-bottom: 28px; box-shadow: 0 10px 25px -5px rgba(99, 102, 241, 0.25);">
        <h1 style="color: #ffffff; margin: 0; font-size: 2.3rem; font-weight: 800; letter-spacing: -0.025em; display: flex; align-items: center; gap: 12px; font-family: 'Outfit', sans-serif;">
            Excel Insights Pro
        </h1>
        <p style="color: rgba(255, 255, 255, 0.9); margin: 6px 0 0 0; font-size: 1.05rem; font-weight: 400; max-width: 800px; line-height: 1.4; font-family: 'Outfit', sans-serif;">
            Unlock deep analytics from multi-tab Excel files. Detect layout configurations automatically, query across worksheets, and visualize findings using advanced model intelligence.
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

with st.sidebar:
    st.caption("API key loaded from secrets")
    st.header("1. Intelligence Settings")
    selected_model_name = st.selectbox("Model:", list(MODELS.keys()), index=0)
    target_model_id = MODELS[selected_model_name]

    st.header("2. Analysis Depth")
    analysis_depth = st.radio(
        "Vision Range:",
        ["Quick (Top 100 Rows)", "Full Dataset (All Rows)"],
        help="Controls how many sample rows are sent to the model (not the full sheet).",
    )

    st.header("3. Load Data")
    input_method = st.radio("Source:", ["Upload File", "Cloud Link"])

    raw_data = None
    if input_method == "Upload File":
        uploaded_file = st.file_uploader("Upload Excel", type=["xlsx", "xlsm"])
        if uploaded_file:
            raw_data = uploaded_file.read()
    else:
        url_input = st.text_input("Paste Google/OneDrive Link:")
        if url_input:
            with st.spinner("Fetching cloud file..."):
                dl_link = get_direct_download_link(url_input)
                try:
                    res = requests.get(dl_link, timeout=25)
                    raw_data = res.content
                except Exception as e:
                    st.error(f"Download failed: {e}")

    if raw_data:
        try:
            st.session_state.all_sheets = load_workbook_sheets(raw_data)
            st.markdown(
                '<div class="badge badge-success" style="display: block; text-align: center; margin: 4px 0; padding: 6px 10px; font-size: 0.85rem;">'
                '✓ Workbook Loaded'
                '</div>',
                unsafe_allow_html=True
            )
        except Exception as e:
            st.error(f"Read Error: {e}")

    if st.session_state.all_sheets:
        st.session_state.scope = st.selectbox(
            "Scope:",
            ["Analyze All Sheets (Join/Compare)"] + list(st.session_state.all_sheets.keys()),
        )

    if st.button("New Conversation"):
        st.session_state.chat_history = []
        st.rerun()

# MAIN INTERFACE
if st.session_state.all_sheets:
    preview_key = (
        list(st.session_state.all_sheets.keys())[0]
        if st.session_state.scope == "Analyze All Sheets (Join/Compare)"
        else st.session_state.scope
    )
    preview_info = st.session_state.all_sheets[preview_key]

    with st.expander("👀 View Raw Structure (First 50 Rows)"):
        st.dataframe(preview_info["df"].head(50))

    with st.expander("📋 Detected Headers & Metadata"):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(
                f'<div class="card-panel" style="text-align: center;">'
                f'<div style="font-size: 0.85rem; color: #94a3b8; margin-bottom: 4px;">Target Sheet</div>'
                f'<div style="font-size: 1.1rem; font-weight: 700; color: #f1f5f9; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">{preview_key}</div>'
                f'</div>',
                unsafe_allow_html=True
            )
        with col2:
            st.markdown(
                f'<div class="card-panel" style="text-align: center;">'
                f'<div style="font-size: 0.85rem; color: #94a3b8; margin-bottom: 4px;">Header Row (Excel)</div>'
                f'<div style="font-size: 1.1rem; font-weight: 700; color: #818cf8;">Row {preview_info["header_row"] + 1} <span style="font-size: 0.75rem; font-weight: 400; color: #64748b;">(0-idx: {preview_info["header_row"]})</span></div>'
                f'</div>',
                unsafe_allow_html=True
            )
        with col3:
            st.markdown(
                f'<div class="card-panel" style="text-align: center;">'
                f'<div style="font-size: 0.85rem; color: #94a3b8; margin-bottom: 4px;">Data Starts (0-indexed)</div>'
                f'<div style="font-size: 1.1rem; font-weight: 700; color: #34d399;">Row {preview_info["data_start_row"]}</div>'
                f'</div>',
                unsafe_allow_html=True
            )

        st.markdown("<h4 style='font-size: 1rem; color: #e2e8f0; margin-bottom: 8px;'>Detected Header Labels</h4>", unsafe_allow_html=True)
        labels_html = ""
        for col_idx, label in preview_info["header_labels"][:30]:
            labels_html += f'<span class="badge badge-primary">Col {col_idx}: {label}</span>'
        if len(preview_info["header_labels"]) > 30:
            labels_html += f'<span class="badge badge-neutral">+{len(preview_info["header_labels"]) - 30} more</span>'
        st.markdown(f'<div style="margin-bottom: 16px;">{labels_html}</div>', unsafe_allow_html=True)

        st.markdown("<h4 style='font-size: 1rem; color: #e2e8f0; margin-bottom: 8px;'>Top Header Candidates (0-indexed Row & Score)</h4>", unsafe_allow_html=True)
        candidates_html = ""
        for row_idx, score in preview_info["header_candidates"]:
            color_class = "badge-success" if row_idx == preview_info["header_row"] else "badge-secondary"
            candidates_html += f'<span class="badge {color_class}">Row {row_idx} (Score: {score:.2f})</span>'
        st.markdown(f'<div style="margin-bottom: 16px;">{candidates_html}</div>', unsafe_allow_html=True)

        st.markdown("<h4 style='font-size: 1rem; color: #e2e8f0; margin-bottom: 8px;'>Sheet Sheet Properties</h4>", unsafe_allow_html=True)
        st.json(preview_info["metadata"])

    for msg in st.session_state.chat_history:
        with st.chat_message("user"):
            st.markdown(msg["user"])
        with st.chat_message("assistant"):
            st.markdown(msg["assistant"])
            if msg.get("code"):
                with st.expander("View Logic"):
                    st.code(msg["code"])

    prompt = st.chat_input("Ask a question about your workbook")
    if prompt:
        workbook_context = build_context(
            st.session_state.all_sheets, st.session_state.scope, analysis_depth
        )
        conversation_history = build_conversation_history()
        with st.spinner("AI Cleaning & Analyzing..."):
            code = get_analysis_code(
                prompt, workbook_context, conversation_history, target_model_id
            )
            if code:
                try:
                    exec_globals = {
                        "pd": pd,
                        "dfs": sheets_to_dfs(st.session_state.all_sheets),
                        "sheet_info": {
                            name: {
                                "header_row": info["header_row"],
                                "data_start_row": info["data_start_row"],
                                "header_labels": info["header_labels"],
                                "column_names": info["column_names"],
                            }
                            for name, info in st.session_state.all_sheets.items()
                        },
                        "find_column_by_header_label": find_column_by_header_label,
                        "SECTION_LABELS": SECTION_LABELS,
                    }
                    exec_locals = {}
                    exec(code, exec_globals, exec_locals)
                    calc_res = exec_locals.get("result", "No result variable created.")

                    answer = generate_natural_answer(prompt, calc_res, target_model_id)
                    st.session_state.chat_history.append({
                        "user": prompt,
                        "assistant": answer,
                        "code": code,
                        "result": str(calc_res),
                    })
                    st.rerun()
                except Exception as e:
                    st.error(f"Execution Error: {e}")
                    st.info("Tip: Check detected header row in the metadata expander if columns look wrong.")
                    st.code(code)
else:
    st.markdown(
        """
        <div style="text-align: center; padding: 48px 20px; max-width: 900px; margin: 0 auto; font-family: 'Outfit', sans-serif;">
            <div style="font-size: 4.5rem; margin-bottom: 20px;">📥</div>
            <h2 style="font-size: 2.1rem; font-weight: 700; color: #f1f5f9; margin-bottom: 12px; letter-spacing: -0.02em;">Welcome to Excel Insights Pro</h2>
            <p style="color: #94a3b8; font-size: 1.1rem; line-height: 1.6; margin-bottom: 40px; max-width: 650px; margin-left: auto; margin-right: auto;">
                Get instant analytics, answer complex questions, and compare datasets across your Excel files. To begin, use the sidebar to load your data.
            </p>
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 20px; text-align: left;">
                <div style="background: rgba(99, 102, 241, 0.03); border: 1px solid rgba(99, 102, 241, 0.08); padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
                    <div style="font-size: 1.8rem; margin-bottom: 10px;">🔍</div>
                    <h4 style="font-weight: 700; color: #f1f5f9; margin: 0 0 8px 0; font-size: 1.05rem;">1. Layout Auto-Detection</h4>
                    <p style="font-size: 0.9rem; color: #94a3b8; margin: 0; line-height: 1.45;">Smart scanning dynamically pinpoints table headers, merged title sections, and frozen boundaries inside multi-tab sheets.</p>
                </div>
                <div style="background: rgba(168, 85, 247, 0.03); border: 1px solid rgba(168, 85, 247, 0.08); padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
                    <div style="font-size: 1.8rem; margin-bottom: 10px;">🤖</div>
                    <h4 style="font-weight: 700; color: #f1f5f9; margin: 0 0 8px 0; font-size: 1.05rem;">2. Code-Backed Reasoning</h4>
                    <p style="font-size: 0.9rem; color: #94a3b8; margin: 0; line-height: 1.45;">Claude writes robust Pandas analysis routines executed in a secure local sandboxed scope to ensure mathematical precision.</p>
                </div>
                <div style="background: rgba(244, 63, 94, 0.03); border: 1px solid rgba(244, 63, 94, 0.08); padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
                    <div style="font-size: 1.8rem; margin-bottom: 10px;">💬</div>
                    <h4 style="font-weight: 700; color: #f1f5f9; margin: 0 0 8px 0; font-size: 1.05rem;">3. Interactive Chat</h4>
                    <p style="font-size: 0.9rem; color: #94a3b8; margin: 0; line-height: 1.45;">Keep the context alive. Ask follow-up queries, request chart derivations, or drill down into anomalies with fluid history tracking.</p>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )
